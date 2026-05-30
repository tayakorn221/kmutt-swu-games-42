# -*- coding: utf-8 -*-
"""สร้างไฟล์ Excel เต็มจาก kmutt_data.json"""
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.load(open("kmutt_data.json", encoding="utf-8"))

ORANGE = "F47B20"   # สีส้ม มจธ.
HEADFILL = PatternFill("solid", fgColor=ORANGE)
WHITEB = Font(bold=True, color="FFFFFF", name="TH Sarabun New", size=14)
CELL = Font(name="TH Sarabun New", size=14)
WINFILL = PatternFill("solid", fgColor="E2F0D9")
LOSEFILL = PatternFill("solid", fgColor="FBE5E5")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADFILL; cell.font = WHITEB; cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def pj(lst, sep=" / "):
    return sep.join(lst)

# ---------- Sheet 1: ตารางแมตช์ ----------
ws = wb.active
ws.title = "ตารางแมตช์"
cols = ["วันเวลา", "ประเภท", "รอบ", "คอร์ท", "นักกีฬา มจธ.", "คู่แข่ง", "สังกัดคู่แข่ง", "ผล", "สกอร์"]
ws.append(cols)
for m in data["matches"]:
    dt = m["datetime"]["th"] or m["datetime"]["iso"]
    kp = pj([p["name"] for p in m["kmutt_players"]])
    opp = pj([o["name"] for o in m["opponents"]])
    oppuni = pj(sorted(set(o["uni"] for o in m["opponents"] if o["uni"])))
    ws.append([dt, m["event"], m["round"], m["court"], kp, opp, oppuni, m["result"], m["score"]])
style_header(ws, len(cols))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(cols) + 1):
        cell = ws.cell(r, c); cell.font = CELL; cell.border = BORDER
        cell.alignment = LEFT if c in (5, 6, 7) else CENTER
    res = ws.cell(r, 8).value
    if res == "ชนะ":
        for c in range(1, len(cols) + 1): ws.cell(r, c).fill = WINFILL
    elif res == "แพ้":
        for c in range(1, len(cols) + 1): ws.cell(r, c).fill = LOSEFILL
autofit(ws, [20, 12, 16, 12, 30, 30, 24, 8, 12])

# ---------- Sheet 2: นักกีฬา + ประเภทที่ลง ----------
ws2 = wb.create_sheet("นักกีฬา มจธ.")
# รวมประเภท+คู่ ต่อคน
byp = defaultdict(list)  # name -> list of (event, partner)
for m in data["matches"]:
    names = [p["name"] for p in m["kmutt_players"]]
    for p in m["kmutt_players"]:
        partner = pj([n for n in names if n != p["name"]]) or "-"
        byp[p["name"]].append((m["event"], partner))
cols2 = ["ลำดับ", "ชื่อนักกีฬา", "ประเภทที่ลงแข่ง", "คู่"]
ws2.append(cols2)
i = 0
seen = {}
for pl in data["players"]:
    nm = pl["name"]
    evs = byp.get(nm, [])
    # unique event->partner
    uniq = {}
    for ev, pa in evs:
        uniq[ev] = pa
    if uniq:
        ev_str = "\n".join(uniq.keys())
        pa_str = "\n".join(uniq.values())
    else:
        ev_str = "(ไม่พบแมตช์)"; pa_str = "-"
    i += 1
    ws2.append([i, nm, ev_str, pa_str])
style_header(ws2, len(cols2))
for r in range(2, ws2.max_row + 1):
    for c in range(1, len(cols2) + 1):
        cell = ws2.cell(r, c); cell.font = CELL; cell.border = BORDER
        cell.alignment = LEFT if c in (2, 3, 4) else CENTER
autofit(ws2, [8, 28, 16, 36])

# ---------- Sheet 3: สรุปผลงาน ----------
ws3 = wb.create_sheet("สรุปผลงาน")
total = len(data["matches"])
win = sum(1 for m in data["matches"] if m["result"] == "ชนะ")
lose = sum(1 for m in data["matches"] if m["result"] == "แพ้")
bye = sum(1 for m in data["matches"] if m["result"] == "บาย")
pending = sum(1 for m in data["matches"] if not m["result"])
ws3.append(["สรุปผลงานทีม มจธ.", ""])
ws3.append(["จำนวนนักกีฬา", len(data["players"])])
ws3.append(["จำนวนแมตช์ทั้งหมด", total])
ws3.append(["บาย (ผ่านเข้ารอบ)", bye])
ws3.append(["ชนะ", win])
ws3.append(["แพ้", lose])
ws3.append(["ยังไม่แข่ง/รอผล", pending])
ws3.append(["", ""])
ws3.append(["ประเภท", "จำนวนแมตช์"])
byev = defaultdict(int)
for m in data["matches"]:
    byev[m["event"]] += 1
for ev, n in sorted(byev.items()):
    ws3.append([ev, n])
for r in range(1, ws3.max_row + 1):
    for c in (1, 2):
        cell = ws3.cell(r, c); cell.font = CELL; cell.border = BORDER; cell.alignment = LEFT
ws3.cell(1, 1).font = Font(bold=True, size=16, name="TH Sarabun New", color="FFFFFF")
ws3.cell(1, 1).fill = HEADFILL
ws3.cell(9, 1).font = WHITEB; ws3.cell(9, 1).fill = HEADFILL
ws3.cell(9, 2).font = WHITEB; ws3.cell(9, 2).fill = HEADFILL
autofit(ws3, [26, 16])

fn = "มจธ_มศวเกมส์42.xlsx"
wb.save(fn)
print("บันทึก", fn)
print(f"แมตช์ {total} | ชนะ {win} | แพ้ {lose} | รอผล {pending}")
